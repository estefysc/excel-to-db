# The import - from openpyxl import * - caused problems witn the open built-in function. See below:
# openpyxl.open is an alias for openpyxl.load_workbook, added in version 3.0. You've imported that alias with from openpyxl import *.
from openpyxl import load_workbook
from helpers.csv_manipulator import create_clean_csv
from helpers.os_utils import search_for_file

# When this function executes, it wil create a csv file for each excel sheet in the specified write_directory
# Returns a list of paths of all the csv files that were created
def excel_to_csv_generator(path, write_directory):
    wb_object = load_workbook(path)
    sheetnames = wb_object.sheetnames

    for sheetname in sheetnames:
        # If the csv file already exists, then yield the path of the csv file
        csv_exists = search_for_file(write_directory, sheetname + ".csv")
        if csv_exists:
            print("The csv file " + sheetname + ".csv already exists in the directory " + write_directory)
            yield sheetname + ".csv"
        # If the csv file does not exist, then create the csv file and yield the path of the csv file
        else:
            print("The csv file " + sheetname + ".csv does not exist in the directory " + write_directory + ". Creating the csv file now...")
            values_list = []
            wb_object.active = wb_object[sheetname]
            current_sheet = wb_object.active
            max_row = current_sheet.max_row
            max_column = current_sheet.max_column

            for i in range(1, max_row + 1):
                row = []
                for j in range(1, max_column + 1):
                    cell_obj = current_sheet.cell(row = i, column = j)
                    row.append(cell_obj.value)
                clean_row = [i for i in row if i is not None]
                values_list.append(clean_row)

            # print("Excel file name: " + sheetname + "\n" + "column amount: " + str(max_column) + "\n" + "row amount: " + str(max_row) + "\n" + "The numbers of columns and rows contain possible empty cells")
            csvFile = create_clean_csv(values_list, sheetname + ".csv", write_directory)
            yield csvFile.name
            # csv_files.append(csv_file.name)
    
    # return csv_files

def excel_to_csv(path, write_directory):
    wb_object = load_workbook(path)
    sheetnames = wb_object.sheetnames
    csv_files_list = []

    for sheetname in sheetnames:
        # If the csv file already exists, then yield the path of the csv file
        csv_exists = search_for_file(write_directory, sheetname + ".csv")
        if csv_exists:
            print("The csv file " + sheetname + ".csv already exists in the directory " + write_directory)
        # If the csv file does not exist, then create the csv file and yield the path of the csv file
        else:
            print("The csv file " + sheetname + ".csv does not exist in the directory " + write_directory + ". Creating the csv file now...")
            values_list = []
            wb_object.active = wb_object[sheetname]
            current_sheet = wb_object.active
            max_row = current_sheet.max_row
            max_column = current_sheet.max_column

            for i in range(1, max_row + 1):
                row = []
                for j in range(1, max_column + 1):
                    cell_obj = current_sheet.cell(row = i, column = j)
                    row.append(cell_obj.value)
                clean_row = [i for i in row if i is not None]
                values_list.append(clean_row)

            # print("Excel file name: " + sheetname + "\n" + "column amount: " + str(max_column) + "\n" + "row amount: " + str(max_row) + "\n" + "The numbers of columns and rows contain possible empty cells")
            csvFile = create_clean_csv(values_list, sheetname + ".csv", write_directory)
            csv_files_list.append(csvFile.name)

    return csv_files_list




    


